"""File content extraction from base64-encoded attachments.

Supports:
  - PDF documents (pypdf)
  - Word documents (.docx via python-docx)
  - Excel spreadsheets (.xlsx via openpyxl)
  - CSV / TSV / plain-text files (chardet for encoding detection)
  - Images (returned as-is for multimodal LLM vision)

All extractors are synchronous and operate on raw bytes.  They are called
from :class:`AttachmentProcessor` which handles base64 decoding.
"""

from __future__ import annotations

import base64
import csv
import io
from typing import Final

import structlog

logger = structlog.get_logger("ai_agent.attachments")

# Hard limit: skip extraction for files larger than this (avoid memory blowup).
_MAX_EXTRACT_BYTES: Final[int] = 10 * 1024 * 1024  # 10 MB

# MIME type → extractor mapping.
_IMAGE_MIMES: frozenset[str] = frozenset(
    {
        "image/jpeg",
        "image/png",
        "image/gif",
        "image/webp",
        "image/svg+xml",
    }
)

_TEXT_MIMES: frozenset[str] = frozenset(
    {
        "text/plain",
        "text/csv",
        "text/tab-separated-values",
        "text/markdown",
        "text/html",
        "application/json",
    }
)


class ExtractionError(Exception):
    """Raised when file extraction fails."""


def extract_pdf(raw: bytes) -> str:
    """Extract text from a PDF file."""
    from pypdf import PdfReader

    reader = PdfReader(io.BytesIO(raw))
    pages: list[str] = []
    for i, page in enumerate(reader.pages):
        text = page.extract_text() or ""
        if text.strip():
            pages.append(f"[Page {i + 1}]\n{text.strip()}")
    return "\n\n".join(pages) if pages else "(PDF contained no extractable text)"


def extract_docx(raw: bytes) -> str:
    """Extract text from a .docx file."""
    from docx import Document

    doc = Document(io.BytesIO(raw))
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    return "\n\n".join(paragraphs) if paragraphs else "(Document contained no text)"


def extract_xlsx(raw: bytes) -> str:
    """Extract text from an .xlsx spreadsheet."""
    from openpyxl import load_workbook  # type: ignore[import-untyped]

    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    sheets_text: list[str] = []
    for name in wb.sheetnames:
        ws = wb[name]
        rows: list[str] = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(cells):
                rows.append("\t".join(cells))
        if rows:
            sheets_text.append(f"[Sheet: {name}]\n" + "\n".join(rows))
    wb.close()
    return "\n\n".join(sheets_text) if sheets_text else "(Spreadsheet contained no data)"


def extract_csv(raw: bytes) -> str:
    """Extract text from a CSV/TSV file with auto-detected encoding."""
    text = _decode_text(raw)
    reader = csv.reader(io.StringIO(text))
    lines: list[str] = []
    for i, row in enumerate(reader):
        lines.append("\t".join(row))
        if i >= 500:  # Cap at 500 rows to avoid prompt bloat.
            lines.append(f"... ({i + 1}+ rows truncated)")
            break
    return "\n".join(lines) if lines else "(CSV contained no data)"


def extract_text(raw: bytes) -> str:
    """Extract text from a plain-text file with auto-detected encoding."""
    return _decode_text(raw)


def _decode_text(raw: bytes) -> str:
    """Decode bytes to str using chardet detection with UTF-8 fallback."""
    try:
        import chardet

        result = chardet.detect(raw)
        encoding = result.get("encoding") or "utf-8"
    except ImportError:
        encoding = "utf-8"
    return raw.decode(encoding, errors="replace")


def extract_attachments(
    files: list[tuple[str, str, bytes]],
) -> tuple[str, list[str]]:
    """Extract text from multiple file attachments.

    Args:
        files: List of ``(name, mime_type, raw_bytes)`` tuples.

    Returns:
        ``(combined_text, image_base64_list)`` — extracted text concatenated,
        and a list of base64 strings for images (for multimodal LLM).
    """
    text_parts: list[str] = []
    image_base64s: list[str] = []

    for name, mime, raw in files:
        if len(raw) > _MAX_EXTRACT_BYTES:
            logger.warning("attachment_too_large", name=name, size=len(raw))
            text_parts.append(f"[{name}] (skipped: file too large for extraction)")
            continue

        try:
            if mime in _IMAGE_MIMES:
                # Images are handled via multimodal LLM, not text extraction.
                # Store raw base64 for the vision API.
                image_base64s.append(base64.b64encode(raw).decode("ascii"))
                text_parts.append(f"[Image: {name} — sent to vision model]")
            elif mime == "application/pdf":
                text_parts.append(f"[{name}]\n{extract_pdf(raw)}")
            elif mime in (
                "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            ):
                text_parts.append(f"[{name}]\n{extract_docx(raw)}")
            elif mime in ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",):
                text_parts.append(f"[{name}]\n{extract_xlsx(raw)}")
            elif mime in ("text/csv", "text/tab-separated-values"):
                text_parts.append(f"[{name}]\n{extract_csv(raw)}")
            elif mime in _TEXT_MIMES or mime.startswith("text/"):
                text_parts.append(f"[{name}]\n{extract_text(raw)}")
            else:
                text_parts.append(f"[{name}] (unsupported file type: {mime})")
        except Exception:
            logger.exception("extraction_failed", name=name, mime=mime)
            text_parts.append(f"[{name}] (extraction failed)")

    return "\n\n".join(text_parts), image_base64s
